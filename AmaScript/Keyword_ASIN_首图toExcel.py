# -*- coding: utf-8 -*-
"""
Created on Thu Mar 18 21:25:20 2021

@author: l1569
"""

import random
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from PIL import Image
# from openpyxl.drawing.image import Image
import os
import urllib
import numpy as np
import requests as req
from io import BytesIO
from skimage import io
import xlwings as xw
from tqdm import tqdm


    # #tqdm (self,  iterable, desc= "Text You want", mininterval=3, initial=50)desc：可以设置进度条前面的文字,mininterval: 设置进度条显示的频率,initial：设置进度条的初始值
    # for i in tqdm(range(1,total_change_num +1),desc= "竞价修改进度", mininterval=5):

# C:/Users/liuhaolin/Desktop/
# C:/Users/l1569/Desktop/
filename = 'C:/Users/liuhaolin/Desktop/' + '烧烤手套类目ASIN' + '.xlsx'
img_path = 'D://Reports//img'
os.makedirs(img_path,exist_ok=True)
by_col = 'B'
to_col = 'R'
img_Num = 49
wb = load_workbook(filename)
ws = wb.active
print('出发！')
for i,Col in enumerate(tqdm(ws[by_col]), start=1):
    # print(i, Col.value)
    if Col.value == 'img':
        continue

    ws.row_dimensions[i].height = 82
    ws.column_dimensions[to_col].width= 16.5
    
    Try_Count = 0
    while Try_Count <=6:
        try:
            time.sleep(random.randint(1,3))
            img_url_read = io.imread(Col.value,plugin='matplotlib')
            break
        except Exception as e:
            Try_Count += 1
            print('%s 出错，出错原因为：%s，第 %s 尝试再次获取'%(Col.value[img_Num:],e,Try_Count))
            time.sleep(5)
            continue

    try:
        height = 100
        width = 128
        IM = Image.fromarray(img_url_read).resize((width,height),Image.ANTIALIAS)
        IM.save(img_path + '/%s'%(Col.value[img_Num:]))
        IMG = openpyxl.drawing.image.Image(img_path + '/%s'%(Col.value[img_Num:]))
        ws.add_image(img=IMG, anchor=to_col + str(i))
    except:
        continue
    
wb.save(filename)
print('ready')


# # C:/Users/liuhaolin/Desktop/
# # C:/Users/l1569/Desktop/
# filename = 'C:/Users/liuhaolin/Desktop/' + 'Competitor_沙滩庭院伞' + '.xlsx'
# img_path = 'D://Reports//img'
# os.makedirs(img_path,exist_ok=True)
# by_col = 'Z'
# to_col = 'AE'
# wb = load_workbook(filename)
# ws = wb.active
# print('出发！')
# for i,Col in enumerate(tqdm(ws[by_col]), start=1):
#     time.sleep(random.random())
#     # print(i, Col.value)
#     try:
#         ws.row_dimensions[i].height = 82
#         ws.column_dimensions[to_col].width= 16.5
        
#         img_url_read = io.imread(Col.value,plugin='matplotlib')
        
#         height = 100
#         width = 128
#         IM = Image.fromarray(img_url_read).resize((width,height),Image.ANTIALIAS)
#         IM.save(img_path + '/%s'%(Col.value[36:]))
#         IMG = openpyxl.drawing.image.Image(img_path + '/%s'%(Col.value[36:]))
#         ws.add_image(img=IMG, anchor=to_col + str(i))
#     except Exception as e:
#         print('%s 出错，出错原因为：%s'%(Col.value[36:],e))
#         continue
    
# wb.save(filename)
# print('ready')



# # 本地文件版本
# from PIL import Image
# filename = 'C:/Users/liuhaolin/Desktop/' + '马桶刷-泵类目ASIN' + '.xlsx'
# img_path = 'D:/Reports/img/'
# img_path2 =  'D:/Reports/img2/'
# by_col = 'B'
# to_col = 'S'
# img_Num = 49
# wb = load_workbook(filename)
# ws = wb.active
# print('出发！')
# for i,Col in enumerate(ws[by_col], start=1):
#     # print(i, Col.value)
#     ws.row_dimensions[i].height = 82
#     ws.column_dimensions[to_col].width= 16.5
#     Img_Read = os.path.join(img_path, Col.value[img_Num:])
#     # print(i, Img_Read)
    
#     try:
#         height = 100
#         width = 250
#         img_url_read = io.imread(Img_Read)
#         IM = Image.fromarray(img_url_read).resize((width,height),Image.ANTIALIAS)
#         IM.save(img_path2 + Col.value[img_Num:])
#         IMG = openpyxl.drawing.image.Image(img_path2 + Col.value[img_Num:])
#         ws.add_image(img=IMG, anchor=to_col + str(i))

#     except:
#         continue
    
# wb.save(filename)
# print('ready')




# File_Name_List = ['C:/Users/liuhaolin/Desktop/img/Competitor_电脑桌.xlsx']
# for filename in File_Name_List:
#     by_col = 'Z'
#     to_col = 'AG'
#     wb = load_workbook(filename)
#     ws = wb.active
#     print('出发！')
#     for i,Col in tqdm(enumerate(ws[by_col], start=1),desc= "抓取进度", mininterval=5):
#         # print(i, Col.value)
#         try:
#             ws.row_dimensions[i].height = 70
#             ws.column_dimensions[to_col].width= 14
            
#             time.sleep(random.randint(1,2.2))
#             img_url_read = io.imread(Col.value,plugin='matplotlib')
            
#             height = 87
#             width = 100
#             IM = Image.fromarray(img_url_read).resize((width,height),Image.ANTIALIAS)
#             IM.save('C:/Users/Administrator/Desktop/img/%s'%(Col.value[36:]))
#             IMG = openpyxl.drawing.image.Image('C:/Users/Administrator/Desktop/img/%s'%(Col.value[36:]))
#             ws.add_image(img=IMG, anchor=to_col + str(i))
#         except:
#             continue
        
#     wb.save(filename)
#     time.sleep(180)
# print('ready')



# filename = 'C:/Users/l1569/Desktop/Competitor_电脑桌+.xlsx'
# by_col = 'Z'
# to_col = 'AG'
# wb = load_workbook(filename)
# ws = wb.active
# for i,Col in enumerate(ws[by_col], start=1):
#     # print(i, Col.value)
#     try:
#         ws.row_dimensions[i].height = 32
#         img_url_read = io.imread(Col.value,plugin='matplotlib')
#         IM = Image.fromarray(img_url_read)
#         IM.save('C:/Users/l1569/Desktop/img/%s.jpg'%(str(i)))
#         Img = Image('C:/Users/l1569/Desktop/img/%s.jpg'%(str(i)))
#         # Img.width = 30
#         # Img.height = 30
#         ws.add_image(img=Img, anchor=to_col + str(i))
#     except:
#         continue
    
# wb.save(filename)

        



# def insert_ing_to_excel(filename,by_col,to_col):
#     wb = load_workbook(filename)
#     ws = wb.active
#     for i,Col in enumerate(ws[by_col], start=1):
#         # print(i, Col.value)
#         ws.row_dimensions[i].height = 32
#         img_url_read = io.imread(Col.value,plugin='matplotlib')
#         width = 30
#         height = 30
#         img_read = Image.fromarray(img_url_read).resize((width, height),Image.ANTIALIAS)
#         ws.add_image(img=img_read,anchor=to_col + str(i))
#     wb.save(filename)

        
# if __name__ == '__main__':
#     insert_ing_to_excel(
#         'C:/Users/l1569/Desktop/Competitor_电脑桌+.xlsx',
#         'Z',
#         'AG')
    

# width = 30
# height = 30
# image = io.imread('https://m.media-amazon.com/images/I/71FaDG3LLnL._AC_UL320_.jpg')#.resize((width, height),Image.ANTIALIAS)
# IM = Image.fromarray(image)
# IM.save('C:/Users/l1569/Desktop/img/1.jpg')
