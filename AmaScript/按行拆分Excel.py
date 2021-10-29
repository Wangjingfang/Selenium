#实现excel表格按行数分拆的功能
import openpyxl

#填入文件名
# excel_name_input=input("填写excel文件名称，限定为.xlsx格式 ")
path = input("请输入要处理的excel文件路径：").replace('"','')
excel_name_open = path
print(excel_name_open)
# excel_name_open=excel_name_input+".xlsx"

# 读取对应excel文件，应加入错误提示，待优化
workbook=openpyxl.load_workbook(filename=excel_name_open)
sheet_origin = workbook.active  #获取活跃的表格

#获取原表格中限定条数据，并复制到新表格

nrows = sheet_origin.max_row  # 最大行数
ncols = sheet_origin.max_column  # 最大列数


#总共需要多少excel

limit=int(input("输入分页数据量 "))
sheets = nrows / limit
if not sheets.is_integer():  #如果不是整除则需要+1
	sheets = int(sheets) + 1


for i in range(1,sheets+1):
	wb =openpyxl.Workbook()
	sheet = wb['Sheet']
	# 写入第一行数据
	for n in range(1,ncols+1):
		sheet.cell(row=1,column=n).value=sheet_origin.cell(row=1,column=n).value
	# 写入范围内数据
	t=2+limit*(i-1)
	num_index=2
	for row_num in range(t,t+limit+1):
		for col_num in range(1,ncols+1) :
			sheet.cell(row=num_index,column=col_num).value=sheet_origin.cell(row=row_num,column=col_num).value
		num_index=num_index+1
	wb.save("{excelname}.xlsx".format(excelname=i))

print('已完成数据拆分')
